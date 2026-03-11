from flask import Flask, render_template, request, jsonify
from datetime import date
import socket
import os
import sys
from database import (
    registrar_consumo_pendiente, 
    obtener_nombres_moviles, 
    obtener_todos_los_skus_para_movimiento,
    obtener_detalles_moviles,
    registrar_consumo_directo,
    inicializar_bd
)
import threading
import json
from config import (
    DB_TYPE, PAQUETES_MATERIALES, MATERIALES_COMPARTIDOS, 
    MOVILES_SANTIAGO, PRODUCTOS_CON_CODIGO_BARRA
)

# Mapeo SKU → Nombre Excel (para mostrar nombres cortos en portal)
SKU_TO_EXCEL_NAME = {
    "1-2-16": "FIBUNHILO",
    "7-1-171": "C_UTP_CAT6",
    "10-1-04": "CONEC_RJ45",
    "5-2-443": "MOLDU",
    "1-4-61": "PLACAS_F_O",
    "1-8-40": "FAJILLA_8",
    "1-8-41": "TAPE",
    "2-5-02": "GRAPAS",
    "2-5-03": "G_C_PARED6",
    "2-7-07": "CALCAMONIA",
    "2-7-11": "COLILLA",
    "4-2-41": "TOALLAS",
    "4-3-18": "CONEC_MECA",
    "4-3-42": "TENSOR_FO",
    "U4-4-633": "HG8247W5",
    "4-4-644": "O_EG8145V5",
    "4-4-654": "O_EG8041X6",
    "4-4-656": "O_EG8041X6",
    "4-4-646": "R_K562E_10",
    "4-4-647": "WIFI_NET",
    "8-1-902": "T_PLAYPRO",
    "8-1-903": "T_PLAY",
    "8-1-904": "E_T_PLAY",
}

if getattr(sys, 'frozen', False):
    # Estamos corriendo en ejecutable (PyInstaller)
    base_dir = sys._MEIPASS
    template_folder = os.path.join(base_dir, 'templates')
    app = Flask(__name__, template_folder=template_folder)
else:
    # Estamos corriendo como script normal
    app = Flask(__name__)

# INICIALIZACIÓN DE BASE DE DATOS (MIGRACIONES)
try:
    print("🔧 Verificando esquema de base de datos...")
    inicializar_bd()
    print("✅ Base de Datos inicializada correctamente.")
except Exception as e:
    print(f"⚠️ Error inicializando BD: {e}")

@app.route('/')
def index():
    status = "OK"
    engine = DB_TYPE
    error_detail = ""
    count_m = 0
    count_p = 0
    
    from config import MYSQL_HOST, MYSQL_USER, MYSQL_PASS, MYSQL_DB
    
    # Validar si faltan variables críticas en la nube
    missing_vars = []
    if DB_TYPE == 'MYSQL':
        if not MYSQL_HOST: missing_vars.append("MYSQL_HOST")
        if not MYSQL_USER: missing_vars.append("MYSQL_USER")
        if not MYSQL_PASS: missing_vars.append("MYSQL_PASSWORD")
        if not MYSQL_DB: missing_vars.append("MYSQL_DATABASE")
    
    if missing_vars:
        status = "CONFIGURACIÓN INCOMPLETA"
        error_detail = f"Faltan variables en Render: {', '.join(missing_vars)}"
    else:
        try:
            moviles = obtener_nombres_moviles()
            productos = obtener_todos_los_skus_para_movimiento()
            # FILTER: Remove SKU 4-4-654 as requested by user
            productos = [p for p in productos if p[1] != '4-4-654']
            
            # MODIFICADO: Usar nombres del Excel en vez de nombres largos de BD
            productos_excel = []
            for nombre_largo, sku, cantidad in productos:
                nombre_excel = SKU_TO_EXCEL_NAME.get(sku, nombre_largo)  # Fallback a nombre largo si no está en mapeo
                productos_excel.append((nombre_excel, sku, cantidad))
            
            details_moviles = obtener_detalles_moviles()
            
            count_m = len(moviles)
            count_p = len(productos_excel)
            
            if count_m == 0 and count_p == 0:
                status = "BASE DE DATOS VACÍA"
                error_detail = f"Conectado a DB: '{MYSQL_DB or 'test'}' en {MYSQL_HOST}. No se encontraron productos ni móviles."
                
        except Exception as e:
            status = "ERROR DE CONEXIÓN"
            error_detail = str(e)
            moviles = []
            productos = []
            details_moviles = {}

    try:
        return render_template('index.html', 
                                 hoy=date.today().isoformat(), 
                                 moviles=moviles if 'moviles' in locals() else [], 
                                 productos=productos_excel if 'productos_excel' in locals() else [],
                                 details_moviles=json.dumps(details_moviles if 'details_moviles' in locals() else {}),
                                 sku_to_excel_name=json.dumps(SKU_TO_EXCEL_NAME),
                                 paquetes=json.dumps(PAQUETES_MATERIALES),
                                 materiales_compartidos=json.dumps(MATERIALES_COMPARTIDOS),
                                 db_status=status,
                                 db_engine=engine,
                                 error_detail=error_detail,
                                 count_m=count_m,
                                 count_p=count_p)
    except Exception as template_err:
        return f"<h1>⚠️ Error de Servidor</h1><p>Estado: {status}</p><p>Detalle: {error_detail}</p><p>Template: {str(template_err)}</p>"

@app.route('/santiago')
def santiago():
    """Página principal del Portal Santiago (Consumo Directo)"""
    status = "OK"
    error_detail = ""
    productos_santiago = []
    
    try:
        # En Santiago Directo, obtenemos TODO el stock de BODEGA
        from config import MYSQL_DB as _MYSQL_DB_S
        target_db = _MYSQL_DB_S  # Forzar DB correcta de Render
        
        # Obtenemos el stock actual de la BD
        raw_stock = obtener_todos_los_skus_para_movimiento(target_db=target_db, sucursal_context='SANTIAGO')
        stock_map = {sku: cantidad for _, sku, cantidad in raw_stock}
        
        # Traemos la lista maestra de config para asegurar que salgan todos (incluso con stock 0)
        from config import PRODUCTOS_INICIALES
        
        for name_excel, sku, _ in PRODUCTOS_INICIALES:
            # Evitar mostrar el router 4-4-654 si anteriormente se pidió filtrar, 
            # pero el usuario ahora pide "toda la lista", así que lo dejamos si está en la maestra.
            cantidad = stock_map.get(sku, 0)
            es_equipo = sku in PRODUCTOS_CON_CODIGO_BARRA
            
            productos_santiago.append({
                "nombre": name_excel,
                "sku": sku,
                "stock": cantidad,
                "es_equipo": es_equipo
            })
            
    except Exception as e:
        status = "ERROR"
        error_detail = str(e)
        logger.error(f"Error cargando portal Santiago: {e}")

    return render_template('santiago.html',
                           hoy=date.today().isoformat(),
                           moviles=MOVILES_SANTIAGO,
                           productos=productos_santiago,
                           sku_to_excel_name=json.dumps(SKU_TO_EXCEL_NAME),
                           db_status=status,
                           error_detail=error_detail)

@app.route('/registrar_santiago', methods=['POST'])
def registrar_santiago_post():
    """Procesa el consumo directo desde el portal de Santiago"""
    data = request.json
    if not data:
        return jsonify({"exito": False, "mensaje": "Sin datos"})

    try:
        exitos = 0
        materiales = data.get('materiales', [])
        movil = data.get('movil')
        tecnico = data.get('tecnico')
        ticket = data.get('contrato')
        fecha = data.get('fecha', date.today().isoformat())

        from database import verificar_seriales_bodega
        from config import MYSQL_DB as _MYSQL_DB

        for item in materiales:
            sku = item['sku']
            cantidad = item.get('cantidad', 1)
            seriales = item.get('seriales', [])
            
            # 1. DETERMINAR CONTEXTO Y DB
            sucursal_ctx = 'SANTIAGO' if movil in MOVILES_SANTIAGO else 'CHIRIQUI'
            # Forzar DB correcta de Render para evitar 'inventario_santiago'
            target_db_ctx = _MYSQL_DB

            # 2. VALIDAR SERIALES/MACs (EQUIPOS)
            if seriales:
                ok_v, msg_v = verificar_seriales_bodega(seriales, sucursal_context=sucursal_ctx, target_db=target_db_ctx)
                if not ok_v:
                    return jsonify({"exito": False, "mensaje": msg_v})

            # 3. REGISTRAR
            exito, msg = registrar_consumo_directo(
                sku=sku,
                cantidad=cantidad,
                movil=movil,
                tecnico=tecnico,
                ticket=ticket,
                fecha_evento=fecha,
                seriales=seriales,
                observaciones=f"Consumo Web Santiago - Ticket {ticket}",
                target_db=target_db_ctx,
                sucursal_context=sucursal_ctx
            )
            
            if not exito:
                raise Exception(f"Error en {sku}: {msg}")
            
            exitos += 1
            
        return jsonify({"exito": True, "mensaje": f"Consumo registrado exitosamente ({exitos} items)."})

    except Exception as e:
        logger.error(f"Error en registrar_santiago: {e}")
        return jsonify({"exito": False, "mensaje": str(e)})

@app.route('/api/validar_serial')
def api_validar_serial():
    """Valida que un serial/MAC exista, esté en BODEGA y corresponda al SKU indicado."""
    serial = request.args.get('serial', '').strip()
    movil = request.args.get('movil', '').strip()
    sku = request.args.get('sku', '').strip()
    
    if not serial or not movil or not sku:
        return jsonify({"exito": False, "mensaje": "Falta serial, móvil o SKU de referencia."})
        
    conn = None
    try:
        from database import get_db_connection, run_query
        from config import MOVILES_SANTIAGO, MYSQL_DB
        
        sucursal = 'SANTIAGO' if movil in MOVILES_SANTIAGO else 'CHIRIQUI'
        # Forzar uso de la DB correcta (variable de entorno MYSQL_DATABASE de Render)
        # Evitar que get_current_db_name() devuelva 'inventario_santiago' por contexto dinámico
        target_db = MYSQL_DB
        
        conn = get_db_connection(target_db=target_db)
        cursor = conn.cursor()
        
        clean_sn = serial.upper()
        
        # Verificar la existencia, ubicación y SKU
        sql = """
            SELECT sku, ubicacion, estado 
            FROM series_registradas 
            WHERE (UPPER(serial_number) = ? OR UPPER(mac_number) = ?)
            AND sucursal = ?
        """
        run_query(cursor, sql, (clean_sn, clean_sn, sucursal))
        result = cursor.fetchone()
        
        if not result:
            return jsonify({"exito": False, "mensaje": f"El serial/MAC {clean_sn} no existe en nuestra base de datos."})
            
        db_sku, db_ubicacion, db_estado = result
        
        if db_ubicacion != 'BODEGA':
            return jsonify({"exito": False, "mensaje": f"El equipo no está en BODEGA (Ubicación actual: {db_ubicacion})."})
            
        if db_sku != sku:
            # Obtener nombre real para mensaje de error más claro si es posible
            run_query(cursor, "SELECT nombre FROM productos WHERE sku = ?", (db_sku,))
            n = cursor.fetchone()
            nombre_err = n[0] if n else "Otro producto"
            return jsonify({"exito": False, "mensaje": f"El serial corresponde a: {nombre_err} (SKU: {db_sku}), pero se esperaba este producto (SKU: {sku})."})
            
        return jsonify({"exito": True, "mensaje": "Validado"})
        
    except Exception as e:
        return jsonify({"exito": False, "mensaje": f"Error verificando: {e}"})
    finally:
        if conn:
            conn.close()

@app.route('/debug/productos')
def debug_productos():
    """Endpoint de diagnóstico para verificar productos en BD"""
    from database import get_db_connection, run_query
    from config import MYSQL_HOST, MYSQL_USER, MYSQL_DB
    
    html = "<h1>🔍 Diagnóstico de Productos</h1>"
    html += f"<p><b>DB_TYPE:</b> {DB_TYPE}</p>"
    html += f"<p><b>MYSQL_HOST:</b> {MYSQL_HOST}</p>"
    html += f"<p><b>MYSQL_DB:</b> {MYSQL_DB}</p>"
    html += f"<p><b>MYSQL_USER:</b> {MYSQL_USER}</p>"
    html += "<hr>"
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Test 1: Total productos
        run_query(cursor, "SELECT COUNT(*) FROM productos")
        total = cursor.fetchone()[0]
        html += f"<p><b>Total productos en tabla:</b> {total}</p>"
        
        # Test 2: Productos únicos
        run_query(cursor, "SELECT COUNT(DISTINCT sku) FROM productos")
        unicos = cursor.fetchone()[0]
        html += f"<p><b>SKUs únicos:</b> {unicos}</p>"
        
        # Test 3: Primeros 10
        run_query(cursor, "SELECT nombre, sku, ubicacion FROM productos LIMIT 10")
        productos = cursor.fetchall()
        html += "<h3>Primeros 10 productos:</h3><ul>"
        for nombre, sku, ub in productos:
            html += f"<li>{sku} | {nombre} | {ub}</li>"
        html += "</ul>"
        
        # Test 4: Función obtener_todos_los_skus_para_movimiento
        from database import obtener_todos_los_skus_para_movimiento
        result = obtener_todos_los_skus_para_movimiento()
        html += f"<h3>obtener_todos_los_skus_para_movimiento():</h3>"
        html += f"<p><b>Retornó:</b> {len(result)} productos</p>"
        if result:
            html += "<ul>"
            for nombre, sku, qty in result[:10]:
                html += f"<li>{sku} | {nombre} | Qty BODEGA: {qty}</li>"
            html += "</ul>"
        
        conn.close()
        
    except Exception as e:
        html += f"<p style='color:red'><b>ERROR:</b> {str(e)}</p>"
        import traceback
        html += f"<pre>{traceback.format_exc()}</pre>"
    
    return html

@app.route('/debug/asignaciones')
def debug_asignaciones():
    """Endpoint de diagnóstico para verificar asignaciones de móviles"""
    from database import get_db_connection, run_query
    from config import MYSQL_HOST, MYSQL_DB
    
    html = "<h1>🚚 Diagnóstico de Asignaciones de Móviles</h1>"
    html += f"<p><b>DB:</b> {MYSQL_DB} @ {MYSQL_HOST}</p>"
    html += "<hr>"
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Paso 1: Verificar estructura de la tabla
        html += "<h2>1. Estructura de la tabla</h2>"
        try:
            run_query(cursor, "DESCRIBE asignacion_moviles")
            columnas = cursor.fetchall()
            html += "<table border='1' cellpadding='5' style='border-collapse:collapse'>"
            html += "<tr><th>Columna</th><th>Tipo</th><th>Null</th><th>Key</th><th>Default</th></tr>"
            col_names = []
            for col in columnas:
                html += f"<tr><td><b>{col[0]}</b></td><td>{col[1]}</td><td>{col[2]}</td><td>{col[3]}</td><td>{col[4]}</td></tr>"
                col_names.append(col[0])
            html += "</table>"
            html += f"<p><b>Columnas encontradas:</b> {', '.join(col_names)}</p>"
        except Exception as e:
            html += f"<p style='color:red'>Error obteniendo estructura: {e}</p>"
            conn.close()
            return html
        
        # Paso 2: Contar registros
        html += "<h2>2. Registros en la tabla</h2>"
        try:
            run_query(cursor, "SELECT COUNT(*) FROM asignacion_moviles")
            total = cursor.fetchone()[0]
            html += f"<p><b>Total registros:</b> {total}</p>"
        except Exception as e:
            html += f"<p style='color:red'>Error contando: {e}</p>"
        
        # Paso 3: Mostrar primeros 10 registros SIN especificar columnas
        html += "<h2>3. Primeros 10 registros (SELECT *)</h2>"
        try:
            run_query(cursor, "SELECT * FROM asignacion_moviles LIMIT 10")
            rows = cursor.fetchall()
            if rows:
                html += "<table border='1' cellpadding='5' style='border-collapse:collapse'>"
                html += "<tr>" + "".join(f"<th>{c}</th>" for c in col_names) + "</tr>"
                for row in rows:
                    html += "<tr>" + "".join(f"<td>{val}</td>" for val in row) + "</tr>"
                html += "</table>"
            else:
                html += "<p>No hay registros</p>"
        except Exception as e:
            html += f"<p style='color:red'>Error obteniendo datos: {e}</p>"
        
        conn.close()
        
    except Exception as e:
        html += f"<p style='color:red'><b>ERROR GENERAL:</b> {str(e)}</p>"
        import traceback
        html += f"<pre>{traceback.format_exc()}</pre>"
    
    return html


@app.route('/api/inventario/<movil>')
def get_inventario_movil(movil):
    """
    API para obtener inventario del técnico con seriales disponibles.
    Retorna JSON con inventario actual del móvil.
    Los MATERIALES_COMPARTIDOS aparecen en AMBOS paquetes (A y B).
    """
    from database import get_db_connection, run_query
    from config import PRODUCTOS_CON_CODIGO_BARRA, MATERIALES_COMPARTIDOS
    
    try:
        # Detectar DB correcta (Chiriquí o Santiago)
        target_db = None
        from config import MOVILES_SANTIAGO
        if movil in MOVILES_SANTIAGO:
            pass
            
        conn = get_db_connection(target_db=target_db)
        cursor = conn.cursor()
        
        # DETERMINAR FILTRO DE SUCURSAL
        sucursal_ctx = 'SANTIAGO' if movil in MOVILES_SANTIAGO else 'CHIRIQUI'

        # Obtener TODAS las asignaciones del móvil (FILTRADO POR SUCURSAL)
        sql_asignacion = """
            SELECT (SELECT p2.nombre FROM productos p2 WHERE p2.sku = a.sku_producto AND p2.sucursal = ? LIMIT 1) as nombre,
                   a.sku_producto, SUM(a.cantidad) as total, COALESCE(a.paquete, 'NINGUNO') as paquete
            FROM asignacion_moviles a
            WHERE a.movil = ? AND a.sucursal = ?
            AND a.cantidad > 0
            GROUP BY a.sku_producto, COALESCE(a.paquete, 'NINGUNO')
        """
        run_query(cursor, sql_asignacion, (sucursal_ctx, movil, sucursal_ctx))
        asignacion_rows = cursor.fetchall()
        
        # Agrupar por SKU para poder calcular totales de compartidos
        # Estructura: {sku: {paquete: cantidad, ...}}
        por_sku = {}
        nombres_sku = {}
        for nombre, sku, cantidad, paquete in asignacion_rows:
            nombres_sku[sku] = nombre
            if sku not in por_sku:
                por_sku[sku] = {}
            por_sku[sku][paquete] = por_sku[sku].get(paquete, 0) + cantidad

        # Construir inventario final
        inventario = []
        
        for nombre, sku, cantidad, paquete in asignacion_rows:
            nombre_final = nombres_sku.get(sku, nombre or sku)
            
            if sku in PRODUCTOS_CON_CODIGO_BARRA:
                # FILTRO CRÍTICO: Buscar seriales SOLO de este paquete específico para este SKU y SUCURSAL
                sql_series = """
                    SELECT serial_number 
                    FROM series_registradas 
                    WHERE sku = ? AND ubicacion = ? AND COALESCE(paquete, 'NINGUNO') = ? AND sucursal = ?
                    ORDER BY serial_number
                """
                pq_query = paquete if paquete else 'NINGUNO'
                run_query(cursor, sql_series, (sku, movil, pq_query, sucursal_ctx))
                seriales = [row[0] for row in cursor.fetchall()]
                
                inventario.append({
                    "sku": sku, "nombre": nombre_final, "paquete": paquete,
                    "seriales": seriales, "cantidad_total": len(seriales),
                    "tiene_series": True, "compartido": (sku in MATERIALES_COMPARTIDOS or paquete == 'PERSONALIZADO')
                })
            else:
                inventario.append({
                    "sku": sku, "nombre": nombre_final, "paquete": paquete,
                    "cantidad_total": cantidad, "tiene_series": False, 
                    "compartido": (sku in MATERIALES_COMPARTIDOS or paquete == 'PERSONALIZADO')
                })
        
        conn.close()
        return jsonify({
            "movil": movil,
            "inventario": inventario,
            "total_productos": len(inventario)
        })
        
    except Exception as e:
        return jsonify({
            "error": str(e),
            "movil": movil,
            "inventario": []
        }), 500


@app.route('/debug')
def debug():
    import os
    from config import MYSQL_HOST, MYSQL_USER, DB_TYPE
    from database import get_db_connection
    
    test_conn = "SIN PROBAR"
    test_error = ""
    
    try:
        conn = get_db_connection()
        conn.close()
        test_conn = "ÉXITO"
    except Exception as e:
        test_conn = "FALLO"
        test_error = str(e)
    
    info = {
        "MODO_ACTUAL": DB_TYPE,
        "MYSQL_HOST": MYSQL_HOST,
        "MYSQL_USER": MYSQL_USER,
        "CONEXION_REAL": test_conn,
        "ERROR_CONEXION": test_error,
        "ENV_EXISTS": os.path.exists('.env'),
        "DIRECTORIO": os.getcwd()
    }
    return jsonify(info)

@app.route('/registrar_bulk', methods=['POST'])
def registrar_bulk():
    data = request.json
    if not data:
        return jsonify({"exito": False, "mensaje": "Sin datos"})

    from database import get_db_connection, run_query
    conn = None
    try:
        # LÓGICA DE ENRUTAMIENTO (NUEVO)
        target_db = None
        from config import MOVILES_SANTIAGO
        movil = data.get('movil', '')
        
        if movil in MOVILES_SANTIAGO:
            sucursal_ctx = 'SANTIAGO'
        else:
            sucursal_ctx = 'CHIRIQUI'

        print(f"[ROUTING] Bulk de {movil} -> Sucursal: {sucursal_ctx} (DB: {target_db or 'Default'})")
            
        conn = get_db_connection(target_db=target_db)
        cursor = conn.cursor()
        
        exitos = 0
        materiales = data.get('materiales', [])
        
        for item in materiales:
            sku = item['sku']
            # Extraer seriales si existen
            seriales = item.get('seriales', [])
            seriales_json = json.dumps(seriales) if seriales else None
            cantidad = item.get('cantidad', len(seriales) if seriales else 0)
            
            # 1. DEDUCCIÓN INMEDIATA DEL STOCK (Para que funcione Offline/PC Apagada)
            try:
                # Importar función de movimiento (asegurar que está disponible)
                from database import registrar_movimiento_gui
                
                # Crear observación
                obs = f"Consumo Web - Ticket: {data.get('contrato')} - Colilla: {data.get('colilla')}"
                
                # Registrar movimiento REAL inmediatamente
                exito_mov, msg_mov = registrar_movimiento_gui(
                    sku=sku,
                    tipo_movimiento='CONSUMO_MOVIL',
                    cantidad_afectada=cantidad,
                    movil_afectado=data['movil'],
                    fecha_evento=data['fecha'],
                    paquete_asignado=data.get('paquete'), # Pass package from web portal
                    observaciones=obs,
                    documento_referencia=data.get('contrato'),
                    existing_conn=conn, # Usar misma conexión
                    sucursal_context=sucursal_ctx
                )
                
                if not exito_mov:
                    raise Exception(f"Fallo al descontar {sku}: {msg_mov}")
                    
            except Exception as e:
                # Si falla la deducción, abortamos todo el bloque
                raise Exception(f"Error procesando {sku}: {str(e)}")

            # 2. CREAR REGISTRO DE AUDITORÍA (AUTO_APROBADO)
            # Esto permite que en la PC se vea el registro, pero marcado como ya procesado
            run_query(cursor, """
                INSERT INTO consumos_pendientes 
                (movil, sku, cantidad, tecnico_nombre, ayudante_nombre, ticket, fecha, colilla, num_contrato, seriales_usados, estado, paquete, sucursal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'AUTO_APROBADO', ?, ?)
            """, (
                data['movil'],
                sku,
                cantidad,
                data['tecnico'],
                data.get('ayudante', ''),
                data['contrato'], # ticket
                data['fecha'],
                data['colilla'],
                data['contrato'],  # num_contrato
                seriales_json,
                data.get('paquete', 'NINGUNO'),
                sucursal_ctx
            ))
            
            # 3. Actualizar ubicación de series a CONSUMIDO (Redundante si registrar_movimiento lo hace, pero seguro)
            if seriales:
                print(f"[WEB] Actualizando {len(seriales)} series a CONSUMIDO para {sku} en {sucursal_ctx}")
                for serial in seriales:
                    run_query(cursor, """
                        UPDATE series_registradas
                        SET ubicacion = 'CONSUMIDO'
                        WHERE serial_number = ? AND sucursal = ?
                    """, (serial, sucursal_ctx))
            
            exitos += 1
        
        conn.commit()
        return jsonify({"exito": True, "mensaje": f"Consumo procesado y descontado exitosamente ({exitos} items)"})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"exito": False, "mensaje": f"Error de base de datos: {str(e)}"})
    finally:
        if conn: conn.close()

@app.route('/auditoria')
def auditoria():
    """Página de Auditoría de Terreno y Retorno"""
    try:
        moviles = obtener_nombres_moviles()
        details_moviles = obtener_detalles_moviles()
    except Exception:
        moviles = []
        details_moviles = {}

    return render_template('auditoria.html',
                           hoy=date.today().isoformat(),
                           moviles=moviles,
                           details_moviles=json.dumps(details_moviles),
                           sku_to_excel_name=json.dumps(SKU_TO_EXCEL_NAME))

@app.route('/api/consumos_dia')
def api_consumos_dia():
    """Devuelve los consumos del día para una móvil y fecha dadas"""
    movil = request.args.get('movil', '')
    fecha = request.args.get('fecha', date.today().isoformat())

    try:
        from database import get_db_connection, run_query
        from config import MOVILES_SANTIAGO

        target_db = None
        pass

        conn = get_db_connection(target_db=target_db)
        cursor = conn.cursor()

        # Traer consumos del día para esa móvil (o todas si no se especifica)
        if movil:
            sql = """
                SELECT id, movil, sku, cantidad, tecnico_nombre, ayudante_nombre,
                       ticket, fecha, colilla, num_contrato, seriales_usados, estado
                FROM consumos_pendientes
                WHERE fecha = ? AND movil = ?
                ORDER BY id DESC
            """
            run_query(cursor, sql, (fecha, movil))
        else:
            sql = """
                SELECT id, movil, sku, cantidad, tecnico_nombre, ayudante_nombre,
                       ticket, fecha, colilla, num_contrato, seriales_usados, estado
                FROM consumos_pendientes
                WHERE fecha = ?
                ORDER BY movil, id DESC
            """
            run_query(cursor, sql, (fecha,))

        rows = cursor.fetchall()
        conn.close()

        consumos = []
        for row in rows:
            id_, movil_r, sku, cantidad, tecnico, ayudante, ticket, fecha_r, colilla, contrato, seriales_json, estado = row
            nombre = SKU_TO_EXCEL_NAME.get(sku, sku)
            seriales = []
            if seriales_json:
                try:
                    seriales = json.loads(seriales_json)
                except Exception:
                    seriales = []
            consumos.append({
                "id": id_,
                "movil": movil_r,
                "sku": sku,
                "nombre": nombre,
                "cantidad": cantidad,
                "tecnico": tecnico,
                "ayudante": ayudante or "",
                "ticket": ticket or "",
                "fecha": str(fecha_r),
                "colilla": colilla or "",
                "contrato": contrato or "",
                "seriales": seriales,
                "estado": estado or ""
            })

        return jsonify({"fecha": fecha, "movil": movil, "consumos": consumos, "total": len(consumos)})

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc(), "consumos": []}), 500

@app.route('/api/comparar_excel', methods=['POST'])
def api_comparar_excel():
    """Recibe un Excel, lo parsea y compara con los consumos de la DB"""
    if 'archivo' not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    archivo = request.files['archivo']
    movil = request.form.get('movil', '')
    fecha = request.form.get('fecha', date.today().isoformat())

    try:
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb.active

        # Leer encabezados de la primera fila
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or '').strip())

        # Leer filas de datos
        filas_excel = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                fila = {headers[i]: str(v or '').strip() for i, v in enumerate(row) if i < len(headers)}
                filas_excel.append(fila)

        # Obtener consumos de la DB para comparar
        from database import get_db_connection, run_query
        from config import MOVILES_SANTIAGO

        target_db = None
        pass

        conn = get_db_connection(target_db=target_db)
        cursor = conn.cursor()

        if movil:
            run_query(cursor, """
                SELECT sku, SUM(cantidad) as total
                FROM consumos_pendientes
                WHERE fecha = ? AND movil = ?
                GROUP BY sku
            """, (fecha, movil))
        else:
            run_query(cursor, """
                SELECT sku, SUM(cantidad) as total
                FROM consumos_pendientes
                WHERE fecha = ?
                GROUP BY sku
            """, (fecha,))

        consumos_db = {row[0]: row[1] for row in cursor.fetchall()}
        conn.close()

        return jsonify({
            "headers": headers,
            "filas": filas_excel,
            "consumos_db": {sku: {"cantidad": qty, "nombre": SKU_TO_EXCEL_NAME.get(sku, sku)}
                            for sku, qty in consumos_db.items()},
            "total_filas": len(filas_excel)
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

@app.route('/api/stock_movil/<movil>')
def api_stock_movil(movil):
    """Retorna el inventario actual de un móvil en formato JSON"""
    try:
        from database import obtener_inventario_movil
        inventario = obtener_inventario_movil(movil)
        return jsonify(inventario)
    except Exception as e:
        print(f"Error en api_stock_movil: {e}")
        return jsonify({}), 500

def start_server():
    # Detectar puerto (Render asigna uno dinámicamente)
    port = int(os.environ.get("PORT", 5000))
    print(f"🚀 Portal Web iniciado en puerto: {port}")
    app.run(host='0.0.0.0', port=port)

def get_local_ip():
    """Obtiene la IP local de la máquina"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def start_server_thread():
    """Inicia el servidor Flask en un thread separado y retorna la IP local"""
    local_ip = get_local_ip()
    
    def run_server():
        try:
            app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
        except Exception as e:
            print(f"Error al iniciar servidor web: {e}")
    
    thread = threading.Thread(target=run_server, daemon=True)
    thread.start()
    
    return local_ip

if __name__ == "__main__":
    start_server()
